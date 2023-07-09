import openpyxl

def equipos_en_pesos(main_variables, valor_usd):
    workbook = openpyxl.load_workbook(main_variables)
    sheet = workbook.active
    cotizacion = sheet.cell(row=2, column=3).value * valor_usd
    workbook.close()

    return cotizacion

def escribir_archivo_excel(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Agregar encabezado
    sheet["A1"] = "ID_Equipo"
    sheet["B1"] = "Categoría"
    sheet["C1"] = "SubCat1"
    sheet["D1"] = "SubCat2"
    sheet["E1"] = "Item"
    sheet["F1"] = "Valor_Pesos"
    sheet["G1"] = "Valor_USD"
    sheet["H1"] = "Valor_Jornada"
    sheet["I1"] = "Stock"

    # Agregar datos de ejemplo
    data = [
        (1, "Rodaje", "Cámara", "", "Canon 5DMIV", 0, 3000, 10000, 1),
        (2, "Rodaje", "Sonido", "", "Zoom H4n", 0, 450, 3750, 1),
        (3, "Postproducción", "Computadora", "", "PC-i5", 0, 1500),

        # Agrega aquí tus datos adicionales
    ]

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)
            sheet.cell(row=row_idx, column=6, value=equipos_en_pesos("mandingues/excels/main_variables.xlsx", row_data[6]))
    
    


    # Guardar los cambios y cerrar el archivo
    workbook.save(file_path)
    workbook.close()

def leer_archivo_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Leer datos del archivo y mostrarlos en consola
    for row in sheet.iter_rows(values_only=True):
        print(row)

    workbook.close()

if __name__ == "__main__":
    file_path = "mandingues/excels/equipos.xlsx"

    # Escribir datos en el archivo
    escribir_archivo_excel(file_path)

    # Leer datos del archivo
    leer_archivo_excel(file_path)
