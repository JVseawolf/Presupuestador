import tkinter as tk
import openpyxl

class GastosFijos:
    def __init__(self, id_gasto, cat="", scat1="", scat2="", scat3="", scat4="", scat5="", total = 0):
        self.id_gasto = id_gasto
        self.cat = cat
        self.scat1 = scat1
        self.scat2 = scat2
        self.scat3 = scat3
        self.scat4 = scat4
        self.scat5 = scat5
        self.total = total
    
class Crew:
    def __init__(self, rol, hs = 0, contratado = False, valor_jornada = 0, hs_jornada = 0, total = 0):
        self.rol = rol
        self.hs = hs
        self.contratado = contratado
        self.valor_jornada = valor_jornada
        self.hs_jornada = hs_jornada
        self.total = total

class Equipos:
    def __init__(self, item, valor_pesos=0, valor_usd=0, valor_jornada=0, q_jornadas=0, stock=0, total=0):
        self.item = item
        self.valor_pesos = valor_pesos
        self.valor_usd = valor_usd
        self.valor_jornada = valor_jornada
        self.q_jornadas = q_jornadas
        self.stock = stock
        self.total = total

def load_fijos():
    datosfijos = []
    workbook = openpyxl.load_workbook(gastos_fijos_path)
    sheet = workbook.active
    for row in sheet.iter_rows(values_only=True, min_row=2, max_col=8):
        gasto_fijo = list(row)
        datosfijos.append(gasto_fijo)
    for gasto_fijo in datosfijos:
        my_gastos_fijos[gasto_fijo[0]] = GastosFijos(gasto_fijo[0])
        my_gastos_fijos[gasto_fijo[0]].cat = gasto_fijo[1]
        my_gastos_fijos[gasto_fijo[0]].scat1 = gasto_fijo[2]
        my_gastos_fijos[gasto_fijo[0]].scat2 = gasto_fijo[3]
        my_gastos_fijos[gasto_fijo[0]].scat3 = gasto_fijo[4]
        my_gastos_fijos[gasto_fijo[0]].scat4 = gasto_fijo[5]
        my_gastos_fijos[gasto_fijo[0]].scat5 = gasto_fijo[6]
        my_gastos_fijos[gasto_fijo[0]].total = gasto_fijo[7]
    
    workbook.close
        

def load_equipos():
    datosequipo = []
    workbook = openpyxl.load_workbook(equipos_path)
    sheet = workbook.active

    for row in sheet.iter_rows(values_only=True, min_row=2):
        filaequipo = list(row)
        datosequipo.append(filaequipo)
    for filaequipo in datosequipo:
        my_equipos[filaequipo[0]] = Equipos(filaequipo[0])
        my_equipos[filaequipo[0]].valor_pesos = filaequipo[1]
        my_equipos[filaequipo[0]].valor_usd = filaequipo[2]
        my_equipos[filaequipo[0]].valor_jornada = filaequipo[3]
        my_equipos[filaequipo[0]].stock = filaequipo[4]
    
    workbook.close

def load_crew():
    datos = []
    workbook = openpyxl.load_workbook("mandingues/excels/horas_hombre.xlsx")
    sheet = workbook.active
    
    for row in sheet.iter_rows(values_only=True, min_row=2):
        fila = list(row)
        datos.append(fila)
    for fila in datos:
        my_crew[fila[0]] = Crew(fila[0])
        my_crew[fila[0]].valor_jornada = fila[1]
        my_crew[fila[0]].hs_jornada = fila[2]
    
    workbook.close()

def get_hs_rentables():
    workbook = openpyxl.load_workbook(main_variables_path)
    sheet = workbook.active
    hs_rentables = sheet["A2"].value
    workbook.close
    return hs_rentables
    
def get_amortizacion():
    total_equipos = 0
    for equipo in my_equipos:
        total_equipos += my_equipos[equipo].valor_pesos * my_equipos[equipo].stock
    return total_equipos / 36

def get_fijos():
    totalfijos = 0
    for fijo in my_gastos_fijos:
        totalfijos += my_gastos_fijos[fijo].total
    return totalfijos

def get_imprevistos():
    imprevistos = 0
    for i in my_gastos_fijos:
        imprevistos += my_gastos_fijos[i].total
    imprevistos += get_amortizacion()
    imprevistos = imprevistos * 0.1
    return imprevistos

def get_valor_hs_estructura():    
    valor_hs_estructura = (get_fijos() + get_amortizacion() + get_imprevistos()) / get_hs_rentables()
    return valor_hs_estructura

def actualizar_horas_estructura():
    global parcial_hs_estructura
    hs_estructura = 0
    for tecnico in my_crew:
        hs_estructura += my_crew[tecnico].hs
    parcial_hs_estructura = hs_estructura * get_valor_hs_estructura()
    return parcial_hs_estructura

def actualizar_presu_hs_hombre():
    global parcial_hs_hombre

    parcial_hs_hombre = 0
    for i in my_crew:
        parcial_hs_hombre += my_crew[i].total

#hs_estructua = get_valor_hs_estructura()

def ventana_presu():

    def set_tercerizado(tecnico):
        my_crew[tecnico].contratado = variables_contratado[tecnico].get()
        print(my_crew[tecnico].contratado)

    def aplicar_hora_celda(tecnico):  
        valor_str = entradas[tecnico].get()
        valor_int = int(valor_str)
        my_crew[tecnico].hs = valor_int
        my_crew[tecnico].total = my_crew[tecnico].hs * my_crew[tecnico].valor_jornada / my_crew[tecnico].hs_jornada
        actualizar_horas_estructura()
        actualizar_presu_hs_hombre()
        actualizar_etiquetas()

    def actualizar_etiquetas():
        for idx, tecnico in enumerate(my_crew):
            etiquetas[idx].config(text=f"{tecnico}: {my_crew[tecnico].hs} hs")
        etiqueta_parcial_hshombre.config(text=f"Hs Hombre: ${parcial_hs_hombre}")
        etiqueta_parcial_estructura.config(text=f"Hs Estructura: ${actualizar_horas_estructura()}")

    def sumar_hora_tecnico(tecnico):
        my_crew[tecnico].hs += 1
        my_crew[tecnico].total = my_crew[tecnico].hs * my_crew[tecnico].valor_jornada / my_crew[tecnico].hs_jornada
        actualizar_horas_estructura()
        actualizar_presu_hs_hombre()
        actualizar_etiquetas()

    def restar_hora_tecnico(tecnico):
        if my_crew[tecnico].hs > 0:
            my_crew[tecnico].hs -= 1
            my_crew[tecnico].total = my_crew[tecnico].hs * my_crew[tecnico].valor_jornada / my_crew[tecnico].hs_jornada
            actualizar_horas_estructura()
            actualizar_presu_hs_hombre()
            actualizar_etiquetas()

    # Crear ventana
    ventana = tk.Tk()
    ventana.title("Presupuesto Horas Equipo Técnico")
    # Centrar ventana en la pantalla
    ancho_ventana = 1200
    alto_ventana = 600
    ancho_pantalla = ventana.winfo_screenwidth()
    alto_pantalla = ventana.winfo_screenheight()
    x = (ancho_pantalla // 2) - (ancho_ventana // 2)
    y = (alto_pantalla // 2) - (alto_ventana // 2)
    ventana.geometry(f"{ancho_ventana}x{alto_ventana}+{x}+{y}")

    variables_contratado = {}
    etiquetas = []
    entradas = {}

    contenedor_hs_hombre = tk.Frame(ventana)
    contenedor_hs_hombre.pack(side="left", fill="both", expand=True)

    contenedor_equipos = tk.Frame(ventana, bg="green")
    contenedor_equipos.pack(side="left", fill="both", expand=True)

    contenedor_variables = tk.Frame(ventana, bg="yellow")
    contenedor_variables.pack(side="left", fill="both", expand=True)

    contenedor_final = tk.Frame(ventana, bg="blue")
    contenedor_final.pack(side="right", fill="both", expand=True)


    for tecnico in my_crew:
        contenedor = tk.Frame(contenedor_hs_hombre)
        contenedor.pack(anchor="w", fill="x")
        etiqueta = tk.Label(contenedor, text=f"{tecnico}: {my_crew[tecnico].hs} hs", font=("Arial", 14))
        etiqueta.pack(side="left")
        etiquetas.append(etiqueta)

        variables_contratado[tecnico] = tk.BooleanVar()
        c1 = tk.Checkbutton(contenedor, text="Tercerizado", variable=variables_contratado[tecnico],
        onvalue=True, offvalue=False, command=lambda tec=tecnico: set_tercerizado(tec))
        c1.pack(anchor="w", side="left")

        boton_aplicar_celda = tk.Button(contenedor, text="Aplicar", command=lambda tec=tecnico: aplicar_hora_celda(tec))
        boton_aplicar_celda.pack(anchor="e", side="right", padx=10)

        celda_definir = tk.Entry(contenedor,  width=4)
        celda_definir.pack(anchor="e", side="right", padx=10)
        celda_definir.insert(0, "0")
        entradas[tecnico] = celda_definir

        boton_agregar = tk.Button(contenedor, text="+", command=lambda tec=tecnico: sumar_hora_tecnico(tec))
        boton_agregar.pack(anchor="e", side="top", padx=10)

        boton_restar = tk.Button(contenedor, text="-", command=lambda tec=tecnico: restar_hora_tecnico(tec))
        boton_restar.pack(anchor="e", side="top", padx=10)

    contenedor = tk.Frame(ventana)
    contenedor.pack(anchor="w", fill="x")
    etiqueta_parcial_hshombre = tk.Label(contenedor_hs_hombre, text=f"Hs Hombre: ${parcial_hs_hombre}")
    etiqueta_parcial_hshombre.pack(anchor="center")
    etiqueta_parcial_estructura = tk.Label(contenedor_hs_hombre, text=f"Hs Estructura: ${parcial_hs_hombre}")
    etiqueta_parcial_estructura.pack(anchor="center")

    ventana.mainloop()

if __name__ == "__main__":
    hs_hombre_path = "mandingues/excels/horas_hombre.xlsx"
    gastos_fijos_path = "mandingues/excels/gastos_fijos.xlsx"
    main_variables_path = "mandingues/excels/main_variables.xlsx"
    equipos_path = "mandingues/excels/equipos.xlsx"

    parcial_hs_hombre = 0
    parcial_hs_estructura = 0

    my_crew = {}
    my_equipos = {}
    my_gastos_fijos = {}

    load_crew()
    load_equipos()
    load_fijos()

    ventana_presu()
