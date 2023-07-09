import openpyxl

def escribir_archivo_excel(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Agregar encabezado
    sheet["A1"] = "Horas_Rentables"
    sheet["B1"] = "Ganancia"
    sheet["C1"] = "Cotización_USD"
    sheet["A2"] = 240
    sheet["B2"] = 0.15
    sheet["C2"] = 493   

    # Guardar los cambios y cerrar el archivo
    workbook.save(file_path)
    workbook.close()


def leer_archivo_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Leer datos del archivo y mostrarlos en consola
    for row in sheet.iter_rows(values_only=True):
        print(row)

    workbook.close()

if __name__ == "__main__":
    file_path = "mandingues/excels/main_variables.xlsx"

    # Escribir datos en el archivo
    escribir_archivo_excel(file_path)

    # Leer datos del archivo
    leer_archivo_excel(file_path)

    
