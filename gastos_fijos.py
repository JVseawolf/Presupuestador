import openpyxl

def calcular_amortizacion(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    total_equipos = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        total_equipos += row[5]
    return total_equipos / 36

def calcular_total_imprevistos(sheet):
    total_imprevistos = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        total_imprevistos += row[7]
    return total_imprevistos * 0.1

def actualizar_total_imprevistos(sheet):
    total_imprevistos = calcular_total_imprevistos(sheet)
    sheet.cell(row=2, column=8, value=total_imprevistos)

def escribir_archivo_excel(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Agregar encabezado
    sheet["A1"] = "ID_Gasto_Fijo"
    sheet["B1"] = "Categoría"
    sheet["C1"] = "SubCat1"
    sheet["D1"] = "SubCat2"
    sheet["E1"] = "SubCat3"
    sheet["F1"] = "SubCat4"
    sheet["G1"] = "SubCat5"
    sheet["H1"] = "Total"

    # Agregar datos de ejemplo
    data = [
        (1, "Imprevistos", "", "", "", "", "", 0),
        (2, "Alquiler", "", "", "", "", "", 1000),
        (3, "Servicios", "Licencias Software", "", "", "", "", 500),
        (4, "Servicios", "Internet", "", "", "", "", 300),
        (5, "Amortización", "", "", "", "", "", calcular_amortizacion("mandingues/excels/equipos.xlsx"))
        # Agrega aquí tus datos adicionales
    ]

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row=row_idx, column=col_idx, value=cell_value)

    # Guardar los cambios y cerrar el archivo
    workbook.save(file_path)
    workbook.close()
    
    # Actualizar el total de imprevistos
    actualizar_total_imprevistos(sheet)

    # Guardar los cambios y cerrar el archivo
    workbook.save(file_path)
    workbook.close()

def leer_archivo_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Leer datos del archivo y mostrarlos en consola
    for row in sheet.iter_rows(values_only=True):
        print(row)

    workbook.close()

if __name__ == "__main__":
    file_path = "mandingues/excels/gastos_fijos.xlsx"

    # Escribir datos en el archivo
    escribir_archivo_excel(file_path)

    # Leer datos del archivo
    leer_archivo_excel(file_path)
