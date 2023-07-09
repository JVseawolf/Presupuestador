import openpyxl

def escribir_archivo_excel(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Agregar encabezado
    sheet["A1"] = "ID_Rol"
    sheet["B1"] = "Rol"
    sheet["C1"] = "Valor_Jornada"
    sheet["D1"] = "Horas_Jornada"
    sheet["E1"] = "Valor_Hora"

    # Agregar datos de ejemplo
    data = [
        (1, "Productor", 13000, 8),
        (2, "Camarógrafo", 27000, 12),
        (3, "Editor", 10000, 10),
        (4, "Sonidista", 10500,3)

        # Agrega aquí tus datos adicionales
    ]

    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, cell_value in enumerate(row_data, start=1):
            sheet.cell(row = row_idx, column = col_idx, value = cell_value)
            sheet.cell(row = row_idx, column = 5, value = row_data[2] / row_data[3])
            
            #Vuelca data al excel




    # Guardar los cambios y cerrar el archivo
    workbook.save(file_path)
    workbook.close()


def leer_archivo_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    # Leer datos del archivo y mostrarlos en consola
    for row in sheet.iter_rows(values_only=True):
        print(row)

    workbook.close()

if __name__ == "__main__":
    file_path = "mandingues/excels/horas_hombre.xlsx"

    # Escribir datos en el archivo
    escribir_archivo_excel(file_path)

    # Leer datos del archivo
    leer_archivo_excel(file_path)
